import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUTPUT = "/Users/nandakumar/Documents/01 Other Projects/suresh/appealdesk/Documents/AppealDesk_FunctionalTestCases.xlsx"

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ─── Colour palette ───────────────────────────────────────────────────────────
C = {
    "navy":       "1E3A5F",
    "navy_light": "2D5491",
    "green":      "065F46",
    "green_mid":  "059669",
    "orange":     "92400E",
    "orange_bg":  "FEF3C7",
    "purple":     "6D28D9",
    "purple_bg":  "EDE9FE",
    "blue_bg":    "DBEAFE",
    "blue_mid":   "1D4ED8",
    "gray_hdr":   "F3F4F6",
    "gray_row":   "F9FAFB",
    "white":      "FFFFFF",
    "pass_bg":    "D1FAE5",
    "fail_bg":    "FEE2E2",
    "blk":        "1A1A2E",
    "mid_gray":   "6B7280",
    "border":     "E5E7EB",
    "section_pl": "EFF6FF",
    "section_pl2":"ECFDF5",
    "section_pl3":"FFF7ED",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="1A1A2E", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def border_thin(color="E5E7EB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium(color="9CA3AF"):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ─── Test case data ───────────────────────────────────────────────────────────
# Each entry: (TC_ID, Section, Module, Test Case Name, Steps, Expected Result, Priority, Role)

PLATFORM_TESTS = [
    # ── LOGIN ──────────────────────────────────────────────────────────────────
    ("PT-001","Authentication","Login","Successful login with valid credentials",
     "1. Open appealdesk.saavigen.ai/login\n2. Enter valid super admin email & password\n3. Click Sign In",
     "User is redirected to Platform Dashboard. Platform sidebar is shown.",
     "High","Super Admin"),

    ("PT-002","Authentication","Login","Login fails with invalid credentials",
     "1. Open login page\n2. Enter wrong email or password\n3. Click Sign In",
     "Error message 'Invalid email or password' is shown. User remains on login page.",
     "High","Super Admin"),

    ("PT-003","Authentication","Login","Login page shows platform logo and name",
     "1. Open login page without logging in",
     "Platform logo (from Platform Settings) and platform name are displayed above the login form.",
     "Medium","Any"),

    ("PT-004","Authentication","Login","Forgot password — reset email sent",
     "1. Click 'Forgot password?' on login page\n2. Enter registered email\n3. Click Send Reset Link",
     "'Check your email' confirmation screen appears. Reset email is received.",
     "High","Any"),

    ("PT-005","Authentication","Login","Forgot password — empty email shows error",
     "1. Click 'Forgot password?'\n2. Leave email blank\n3. Click Send Reset Link",
     "Validation error 'Email is required' appears. No email is sent.",
     "Medium","Any"),

    # ── DASHBOARD ──────────────────────────────────────────────────────────────
    ("PT-006","Platform Dashboard","Dashboard","Dashboard loads with correct stats",
     "1. Log in as Super Admin\n2. Navigate to Dashboard",
     "Dashboard shows total service providers count and key platform metrics.",
     "High","Super Admin"),

    # ── SERVICE PROVIDERS ──────────────────────────────────────────────────────
    ("PT-007","Service Providers","Providers List","Service provider list loads",
     "1. Navigate to Service Providers in sidebar",
     "All registered service providers are listed with name, status, and created date.",
     "High","Super Admin"),

    ("PT-008","Service Providers","Create Provider","Create a new service provider",
     "1. Click 'New Provider'\n2. Fill required fields: Name, type\n3. Click Save",
     "New service provider is created and appears in the list. Success message shown.",
     "High","Super Admin"),

    ("PT-009","Service Providers","Create Provider","Required field validation on create",
     "1. Click 'New Provider'\n2. Leave Name empty\n3. Click Save",
     "Validation error shown. Provider is not created.",
     "Medium","Super Admin"),

    ("PT-010","Service Providers","Edit Provider","Edit an existing service provider",
     "1. Click on a provider from the list\n2. Edit name or other fields\n3. Save",
     "Provider details are updated. Changes persist after page refresh.",
     "High","Super Admin"),

    ("PT-011","Service Providers","Provider Detail","View provider detail page",
     "1. Click on a service provider name in the list",
     "Detail page opens showing provider info, PAN, GST, contact details, and associated clients.",
     "Medium","Super Admin"),

    ("PT-012","Service Providers","Logo Upload","Upload logo for service provider",
     "1. Open provider detail / edit\n2. Upload a PNG/JPG image\n3. Save",
     "Logo is uploaded and displayed on the provider record.",
     "Medium","Super Admin"),

    ("PT-013","Service Providers","Deactivate Provider","Deactivate a service provider",
     "1. Open provider detail\n2. Toggle active status to inactive\n3. Confirm",
     "Provider status changes to inactive. SP users for that provider can no longer log in.",
     "High","Super Admin"),

    # ── PLATFORM ADMINS ────────────────────────────────────────────────────────
    ("PT-014","Admins","Admin List","Platform admins list loads",
     "1. Navigate to Admins in sidebar",
     "All platform admin users are shown in a list.",
     "High","Super Admin"),

    ("PT-015","Admins","Create Admin","Create a new platform admin",
     "1. Click 'New Admin'\n2. Enter name, email, password\n3. Save",
     "New admin account is created. User receives credentials and can log in.",
     "High","Super Admin"),

    ("PT-016","Admins","Deactivate Admin","Deactivate a platform admin",
     "1. Select an admin\n2. Toggle to inactive\n3. Confirm",
     "Admin account is deactivated. That user can no longer log in.",
     "High","Super Admin"),

    # ── MASTER RECORDS ─────────────────────────────────────────────────────────
    ("PT-017","Master Records","Masters","Master records list loads",
     "1. Navigate to Master Records in sidebar",
     "Reference data (authority types, proceeding types, jurisdictions) are listed.",
     "Medium","Super Admin"),

    ("PT-018","Master Records","Masters","Add a new master record",
     "1. Click Add on the relevant master section\n2. Enter value\n3. Save",
     "New record appears in the list and is available in SP portal dropdowns.",
     "Medium","Super Admin"),

    ("PT-019","Master Records","Masters","Edit an existing master record",
     "1. Click edit on a master record\n2. Change the value\n3. Save",
     "Updated value appears in list and in SP portal dropdowns.",
     "Medium","Super Admin"),

    # ── PLATFORM SETTINGS ──────────────────────────────────────────────────────
    ("PT-020","Settings","Branding","Upload platform logo",
     "1. Go to Settings\n2. Click 'Change Logo'\n3. Select a JPG/PNG file (≤2MB)\n4. Save",
     "Logo is uploaded and displayed in the sidebar and on the login page.",
     "High","Super Admin"),

    ("PT-021","Settings","Branding","Update platform name",
     "1. Go to Settings\n2. Change Platform Name field\n3. Save",
     "Updated name is reflected in sidebar and login page title.",
     "High","Super Admin"),

    ("PT-022","Settings","Branding","Update support email",
     "1. Go to Settings\n2. Enter support email\n3. Save",
     "Support email appears on the login page as a clickable mailto link.",
     "Medium","Super Admin"),

    ("PT-023","Settings","Branding","Remove platform logo",
     "1. Go to Settings\n2. Click 'Remove logo'\n3. Save",
     "Logo is removed. Login page and sidebar show default icon instead.",
     "Medium","Super Admin"),

    ("PT-024","Settings","Branding","Platform name is required — save without name",
     "1. Go to Settings\n2. Clear Platform Name\n3. Click Save",
     "Validation error shown. Settings are not saved.",
     "Medium","Super Admin"),

    ("PT-025","Settings","Account","Update super admin password",
     "1. Go to Settings → Account section\n2. Enter current password and new password\n3. Save",
     "Password is updated. Old password no longer works for login.",
     "High","Super Admin"),
]

SP_TESTS = [
    # ── AUTHENTICATION ─────────────────────────────────────────────────────────
    ("SP-001","Authentication","Login","SP Admin login",
     "1. Go to login page\n2. Enter SP Admin credentials\n3. Click Sign In",
     "SP Admin is redirected to the SP Dashboard with full sidebar navigation.",
     "High","SP Admin"),

    ("SP-002","Authentication","Login","SP Staff login",
     "1. Go to login page\n2. Enter SP Staff credentials\n3. Click Sign In",
     "SP Staff is redirected to SP Dashboard. Sidebar shows permitted sections.",
     "High","SP Staff"),

    ("SP-003","Authentication","Login","Deactivated user cannot login",
     "1. Use credentials of a deactivated user\n2. Sign In",
     "Error or deactivated message shown. Access is denied.",
     "High","Any"),

    ("SP-004","Authentication","Logout","Logout clears session",
     "1. Click Logout in sidebar\n2. Try to access /appeals directly",
     "User is redirected to login page. Session is cleared.",
     "High","Any"),

    # ── DASHBOARD ──────────────────────────────────────────────────────────────
    ("SP-005","Dashboard","Overview","Dashboard loads for SP Admin",
     "1. Log in as SP Admin\n2. View Dashboard",
     "Stat cards show: Total Appeals, Total Clients, Team count. Importance tiles, outcome tiles, upcoming deadlines, recent activity are all visible.",
     "High","SP Admin"),

    ("SP-006","Dashboard","Overview","Dashboard loads for SP Staff",
     "1. Log in as SP Staff\n2. View Dashboard",
     "Same as SP Admin dashboard. Full stats visible.",
     "High","SP Staff"),

    ("SP-007","Dashboard","Upcoming Deadlines","Upcoming deadlines show correct appeals",
     "1. View Dashboard\n2. Check Upcoming Deadlines panel",
     "Appeals with to_be_completed_by within next 30 days are listed. Days remaining shown in colour (red <7 days, orange <15, green otherwise).",
     "High","SP Admin, SP Staff"),

    ("SP-008","Dashboard","Recent Activity","Recent activity shows latest events",
     "1. View Dashboard\n2. Check Recent Activity panel",
     "Last 5 events (across all appeals) are listed with category and date.",
     "Medium","SP Admin, SP Staff"),

    # ── APPEALS LIST ───────────────────────────────────────────────────────────
    ("SP-009","Appeals","List","Appeals list loads",
     "1. Navigate to Appeals",
     "All appeals for the SP are shown in a table with client name, FY, AY, act, importance, outcome, deadline, assigned to.",
     "High","SP Admin, SP Staff"),

    ("SP-010","Appeals","List","Filter appeals by client",
     "1. On Appeals list, select a client from the Client filter dropdown",
     "Only appeals belonging to the selected client are displayed.",
     "High","SP Admin, SP Staff"),

    ("SP-011","Appeals","List","Filter appeals by financial year",
     "1. Select a financial year from the FY filter",
     "Only appeals matching that financial year are shown.",
     "High","SP Admin, SP Staff"),

    ("SP-012","Appeals","List","Filter appeals by importance",
     "1. Select importance (Critical / High / Medium / Low) from filter",
     "List is filtered to matching importance level only.",
     "Medium","SP Admin, SP Staff"),

    ("SP-013","Appeals","List","Filter appeals by outcome",
     "1. Select outcome (Favourable / Doubtful / Unfavourable) from filter",
     "List is filtered to matching outcome only.",
     "Medium","SP Admin, SP Staff"),

    ("SP-014","Appeals","List","Filter appeals by assigned user",
     "1. Select a staff member from Assigned To filter",
     "Only appeals assigned to that staff member are shown.",
     "Medium","SP Admin, SP Staff"),

    ("SP-015","Appeals","List","Multiple filters applied together",
     "1. Select Client + Importance filters simultaneously",
     "Results satisfy all selected filter conditions simultaneously.",
     "High","SP Admin, SP Staff"),

    ("SP-016","Appeals","List","Clear all filters",
     "1. Apply filters\n2. Click Clear / Reset filters",
     "All filters are reset and full appeal list is restored.",
     "Medium","SP Admin, SP Staff"),

    # ── CREATE APPEAL ──────────────────────────────────────────────────────────
    ("SP-017","Appeals","Create","Create a new appeal with proceeding",
     "1. Click 'New Appeal'\n2. Select client org\n3. Fill FY, AY, Act/Regulation\n4. Fill proceeding details\n5. Save",
     "Appeal is created and appears in the appeals list. Appeal detail page opens.",
     "High","SP Admin, SP Staff"),

    ("SP-018","Appeals","Create","Client org is required on create",
     "1. Click 'New Appeal'\n2. Leave client org blank\n3. Save",
     "Validation error. Appeal is not created.",
     "High","SP Admin, SP Staff"),

    ("SP-019","Appeals","Create","Financial year format",
     "1. Create appeal\n2. Enter FY as '2023-24'\n3. Save",
     "FY is stored and displayed correctly as entered.",
     "Medium","SP Admin, SP Staff"),

    # ── APPEAL DETAIL ──────────────────────────────────────────────────────────
    ("SP-020","Appeals","Detail","Appeal detail page loads correctly",
     "1. Click on an appeal from the list",
     "Detail page shows: appeal header (client, FY, AY, Act), all proceedings, events timeline per proceeding, and documents section.",
     "High","SP Admin, SP Staff"),

    ("SP-021","Appeals","Detail","Client name is displayed on detail page",
     "1. Open appeal detail",
     "Client organisation name is shown in the appeal header. Not blank.",
     "High","SP Admin, SP Staff"),

    ("SP-022","Appeals","Detail","Assigned To is displayed on proceeding",
     "1. Open appeal detail\n2. View proceeding section",
     "Assigned staff member name is shown on the proceeding card. Not blank.",
     "High","SP Admin, SP Staff"),

    ("SP-023","Appeals","Edit","Edit appeal details",
     "1. Open appeal detail\n2. Click Edit Appeal\n3. Change FY or client\n4. Save",
     "Appeal is updated. Changes reflected on detail page immediately.",
     "High","SP Admin, SP Staff"),

    ("SP-024","Appeals","Delete","Delete appeal with confirmation",
     "1. Open appeal detail\n2. Click Delete Appeal\n3. Confirm in modal",
     "Appeal and all its proceedings, events, and documents are deleted. User is redirected to appeals list and the deleted appeal no longer appears.",
     "High","SP Admin, SP Staff"),

    ("SP-025","Appeals","Delete","Cancel appeal delete",
     "1. Click Delete Appeal\n2. Click Cancel in confirmation modal",
     "Modal closes. Appeal is NOT deleted. Detail page remains.",
     "Medium","SP Admin, SP Staff"),

    # ── PROCEEDINGS ────────────────────────────────────────────────────────────
    ("SP-026","Appeals","Proceedings","Add a new proceeding to existing appeal",
     "1. Open appeal detail\n2. Click 'Add Proceeding'\n3. Fill proceeding details\n4. Save",
     "New proceeding is added and appears under the appeal detail.",
     "High","SP Admin, SP Staff"),

    ("SP-027","Appeals","Proceedings","Edit an existing proceeding",
     "1. Open appeal detail\n2. Click Edit on a proceeding\n3. Change importance or assigned user\n4. Save",
     "Proceeding is updated. Changes visible immediately on detail page.",
     "High","SP Admin, SP Staff"),

    ("SP-028","Appeals","Proceedings","Importance level displayed with correct colour",
     "1. View a proceeding with importance set to 'Critical'",
     "Critical is shown in red, High in orange, Medium in yellow, Low in green (or configured colour coding).",
     "Medium","SP Admin, SP Staff"),

    # ── EVENTS ─────────────────────────────────────────────────────────────────
    ("SP-029","Appeals","Events","Add an event to a proceeding",
     "1. Open appeal detail\n2. On a proceeding, click 'Add Event'\n3. Select category (e.g. Personal Hearing)\n4. Fill hearing date and details\n5. Save",
     "Event is added to the proceeding timeline and shown in chronological order.",
     "High","SP Admin, SP Staff"),

    ("SP-030","Appeals","Events","Quick View an event",
     "1. On proceeding events list, click 'Quick View' on an event",
     "A read-only modal opens showing all event details (category, date, all detail fields).",
     "High","SP Admin, SP Staff"),

    ("SP-031","Appeals","Events","Edit an event",
     "1. Click 'Edit' on an event in the proceeding timeline\n2. Change date or description\n3. Save",
     "Event is updated. Updated values shown on the timeline.",
     "High","SP Admin, SP Staff"),

    ("SP-032","Appeals","Events","Delete an event with confirmation",
     "1. Click 'Delete' on an event\n2. Confirm in modal",
     "Event is permanently deleted from the proceeding timeline.",
     "High","SP Admin, SP Staff"),

    ("SP-033","Appeals","Events","Cancel event delete",
     "1. Click 'Delete' on an event\n2. Click Cancel",
     "Modal closes. Event is NOT deleted.",
     "Medium","SP Admin, SP Staff"),

    ("SP-034","Appeals","Events","All event categories available in dropdown",
     "1. Click 'Add Event'\n2. Open category dropdown",
     "All 9 event types are listed: Notice from Authority, Response to Notice, Adjournment Request, Personal Hearing, Virtual Hearing, Personal Follow-up, Assessment Order, Notice of Penalty, Penalty Order.",
     "Medium","SP Admin, SP Staff"),

    ("SP-035","Appeals","Events","Event-specific fields shown based on category",
     "1. Add Event\n2. Select 'Personal Hearing' category",
     "Hearing date, location, attendees and outcome fields appear, relevant to personal hearing.",
     "High","SP Admin, SP Staff"),

    # ── DOCUMENTS (on Appeal) ──────────────────────────────────────────────────
    ("SP-036","Appeals","Documents","Upload a document to an appeal",
     "1. Open appeal detail\n2. Scroll to Documents section\n3. Click Upload\n4. Select a PDF file\n5. Confirm upload",
     "Document is uploaded and appears in the documents list with filename, size, and upload date.",
     "High","SP Admin, SP Staff"),

    ("SP-037","Appeals","Documents","Download a document",
     "1. On appeal detail documents section\n2. Click Download on a file",
     "File downloads to local machine.",
     "High","SP Admin, SP Staff, Client"),

    ("SP-038","Appeals","Documents","Delete a document with confirmation",
     "1. Click Delete on a document in the list\n2. Confirm",
     "Document is removed from storage and no longer listed.",
     "High","SP Admin, SP Staff"),

    # ── CLIENTS ────────────────────────────────────────────────────────────────
    ("SP-039","Clients","List","Clients list loads",
     "1. Navigate to Clients",
     "All client organisations are listed with name, PAN, and status.",
     "High","SP Admin, SP Staff"),

    ("SP-040","Clients","Create","Create a new client organisation",
     "1. Click 'New Client'\n2. Enter organisation name (required)\n3. Optionally add PAN, TAN, GST, address\n4. Save",
     "New client is created and appears in the clients list.",
     "High","SP Admin"),

    ("SP-041","Clients","Create","Client name is required",
     "1. Click 'New Client'\n2. Leave name blank\n3. Save",
     "Validation error shown. Client not created.",
     "High","SP Admin"),

    ("SP-042","Clients","Edit","Edit client details",
     "1. Open a client\n2. Edit PAN or address\n3. Save",
     "Client is updated. Changes visible on client detail.",
     "High","SP Admin"),

    ("SP-043","Clients","Detail","Client detail page shows related appeals",
     "1. Click on a client from the list",
     "Client detail page shows all appeals associated with that client.",
     "High","SP Admin, SP Staff"),

    ("SP-044","Clients","Staff","Add a contact staff member to client",
     "1. Open client detail\n2. Add a contact person (name, designation, mobile)\n3. Save",
     "Contact staff is saved and shown on the client record.",
     "Medium","SP Admin"),

    # ── USERS ──────────────────────────────────────────────────────────────────
    ("SP-045","Users","List","Users list loads with all user types",
     "1. Navigate to Users as SP Admin",
     "List shows SP Admin, SP Staff, and Client users grouped or filterable. Org name shown per user.",
     "High","SP Admin"),

    ("SP-046","Users","Create","Create SP Admin user",
     "1. Click 'New User'\n2. Fill name, email, password\n3. Select Role: SP Admin\n4. Save",
     "SP Admin user is created. User can log in with provided credentials.",
     "High","SP Admin"),

    ("SP-047","Users","Create","Create SP Staff user",
     "1. Click 'New User'\n2. Fill name, email, password\n3. Role: SP Staff\n4. Add department, designation\n5. Save",
     "Staff user is created and visible in Users list.",
     "High","SP Admin"),

    ("SP-048","Users","Create","Create Client user",
     "1. Click 'New User'\n2. Fill name, email, password\n3. Role: Client\n4. Select client org\n5. Save",
     "Client user is created under the correct organisation.",
     "High","SP Admin"),

    ("SP-049","Users","Create","Email must be unique",
     "1. Create a user with an email already in use",
     "Error shown: email already registered. User not created.",
     "High","SP Admin"),

    ("SP-050","Users","Create","Password is required on create",
     "1. Create user\n2. Leave password blank\n3. Save",
     "Validation error. User not created.",
     "High","SP Admin"),

    ("SP-051","Users","Create","SP staff fields shown for SP roles only",
     "1. Create user\n2. Select Role: SP Staff\n3. Check Employment section",
     "Employment, Address, Identity (PAN/Aadhar) sections are visible.",
     "Medium","SP Admin"),

    ("SP-052","Users","Create","SP staff fields hidden for client role",
     "1. Create user\n2. Select Role: Client",
     "Employment, Address, Identity sections are hidden. Only basic info fields shown.",
     "Medium","SP Admin"),

    ("SP-053","Users","Create","Mobile with country code captured",
     "1. Create user\n2. Select country code (e.g. +91)\n3. Enter mobile number",
     "Mobile is saved with country code prefix.",
     "Medium","SP Admin"),

    ("SP-054","Users","Activate/Deactivate","Deactivate a user",
     "1. Open user record\n2. Toggle to inactive\n3. Confirm",
     "User is deactivated. Attempting to login shows deactivated message.",
     "High","SP Admin"),

    ("SP-055","Users","Activate/Deactivate","Reactivate a user",
     "1. Open inactive user\n2. Toggle to active",
     "User is reactivated and can log in again.",
     "High","SP Admin"),

    ("SP-056","Users","Access Control","SP Staff cannot access Users section",
     "1. Log in as SP Staff\n2. Try to navigate to /users",
     "User is redirected or shown 'Unauthorised'. Users section not in sidebar.",
     "High","SP Staff"),

    # ── DOCUMENTS (global) ─────────────────────────────────────────────────────
    ("SP-057","Documents","List","Documents page loads all documents",
     "1. Navigate to Documents",
     "All uploaded documents across all appeals are shown with filename, appeal, client, upload date.",
     "High","SP Admin, SP Staff"),

    ("SP-058","Documents","Filter","Filter documents by client",
     "1. On Documents page, select a client filter",
     "Only documents from that client's appeals are shown.",
     "Medium","SP Admin, SP Staff"),

    ("SP-059","Documents","Download","Download a document from Documents page",
     "1. Click Download on a document row",
     "File downloads to local machine.",
     "High","SP Admin, SP Staff"),

    ("SP-060","Documents","Delete","Delete a document from Documents page",
     "1. Click Delete on a document\n2. Confirm",
     "Document is removed from storage and disappears from list.",
     "High","SP Admin"),

    # ── AUDIT LOGS ─────────────────────────────────────────────────────────────
    ("SP-061","Audit Logs","List","Audit logs load for SP Admin",
     "1. Log in as SP Admin\n2. Navigate to Logs",
     "Chronological list of all logged actions: who, what entity, what action, timestamp.",
     "High","SP Admin"),

    ("SP-062","Audit Logs","Access","SP Staff cannot access Audit Logs",
     "1. Log in as SP Staff\n2. Try to navigate to /logs",
     "'Unauthorised' or redirect. Logs link not shown in sidebar.",
     "High","SP Staff"),

    ("SP-063","Audit Logs","Filter","Filter logs by action type",
     "1. On Logs page, select filter: 'create' or 'delete'",
     "Only log entries matching that action type are shown.",
     "Medium","SP Admin"),

    ("SP-064","Audit Logs","Filter","Filter logs by date range",
     "1. Set from-date and to-date on Logs page",
     "Only log entries within that date range are shown.",
     "Medium","SP Admin"),

    ("SP-065","Audit Logs","Logging","Creating an appeal logs an entry",
     "1. Create a new appeal\n2. Go to Audit Logs",
     "A 'create' entry for 'appeal' entity appears in the logs with actor name and timestamp.",
     "High","SP Admin"),

    ("SP-066","Audit Logs","Logging","Deleting an appeal logs an entry",
     "1. Delete an appeal\n2. Go to Audit Logs",
     "A 'delete' entry for 'appeal' entity appears in the logs.",
     "High","SP Admin"),

    # ── SETTINGS ───────────────────────────────────────────────────────────────
    ("SP-067","Settings","Branding","Upload SP logo in settings",
     "1. Go to Settings\n2. Upload a JPG/PNG logo\n3. Save",
     "Logo appears in sidebar and is visible on the SP portal.",
     "High","SP Admin"),

    ("SP-068","Settings","Branding","Update SP organisation name",
     "1. Go to Settings\n2. Change organisation name\n3. Save",
     "Updated name shown in sidebar.",
     "Medium","SP Admin"),

    ("SP-069","Settings","Access","SP Staff can view settings but not edit branding",
     "1. Log in as SP Staff\n2. Go to Settings",
     "Settings page is visible but edit/save controls for branding are hidden or disabled.",
     "Medium","SP Staff"),

    ("SP-070","Settings","Account","Update own password",
     "1. Go to Settings → My Account\n2. Enter current and new password\n3. Save",
     "Password updated successfully. Old password no longer works.",
     "High","SP Admin, SP Staff"),

    # ── MASTER RECORDS (SP) ────────────────────────────────────────────────────
    ("SP-071","Master Records","List","Master records visible in SP portal",
     "1. Navigate to Master Records",
     "Reference data (authority types, proceeding types, etc.) are listed.",
     "Medium","SP Admin, SP Staff"),
]

CLIENT_TESTS = [
    # ── LOGIN ──────────────────────────────────────────────────────────────────
    ("CL-001","Authentication","Login","Client user successful login",
     "1. Go to login page\n2. Enter client user credentials\n3. Sign In",
     "Client user is redirected to the SP Dashboard with restricted sidebar (Dashboard, Appeals, Documents only).",
     "High","Client"),

    ("CL-002","Authentication","Login","Client user sees restricted sidebar",
     "1. Log in as Client user\n2. Check sidebar navigation",
     "Sidebar shows only: Dashboard, Appeals, Documents. No Users, Clients, Logs, or Settings links.",
     "High","Client"),

    # ── DASHBOARD ──────────────────────────────────────────────────────────────
    ("CL-003","Dashboard","Overview","Dashboard shows only own org's data",
     "1. Log in as Client user\n2. View Dashboard",
     "Stat cards reflect only appeals belonging to the client's own organisation. No other clients' data visible.",
     "High","Client"),

    ("CL-004","Dashboard","Overview","Upcoming deadlines show own org's appeals only",
     "1. View Dashboard as Client user\n2. Check Upcoming Deadlines",
     "Only deadlines for the client's own appeals are listed.",
     "High","Client"),

    # ── APPEALS ────────────────────────────────────────────────────────────────
    ("CL-005","Appeals","List","Client sees only own organisation's appeals",
     "1. Log in as Client user\n2. Navigate to Appeals",
     "Appeals list shows only appeals where client_org_id matches the client user's organisation. No other client's appeals visible.",
     "High","Client"),

    ("CL-006","Appeals","List","Client cannot see Create Appeal button",
     "1. Navigate to Appeals as Client user",
     "'New Appeal' button is not shown. Client cannot initiate an appeal.",
     "High","Client"),

    ("CL-007","Appeals","Detail","Client can view appeal detail",
     "1. Click on an appeal in the list",
     "Appeal detail page opens showing proceedings, events, and documents. All data is read-only.",
     "High","Client"),

    ("CL-008","Appeals","Detail","Client cannot edit appeal",
     "1. Open appeal detail as Client user",
     "No Edit Appeal button is shown. Form fields are read-only or not present.",
     "High","Client"),

    ("CL-009","Appeals","Detail","Client cannot delete appeal",
     "1. Open appeal detail as Client user",
     "No Delete Appeal button is visible.",
     "High","Client"),

    ("CL-010","Appeals","Events","Client can view events in proceeding timeline",
     "1. Open appeal detail as Client user\n2. View proceeding events",
     "Events timeline is visible with Quick View option. Edit and Delete buttons are NOT shown.",
     "High","Client"),

    ("CL-011","Appeals","Events","Client cannot add events",
     "1. Open appeal detail as Client user\n2. Look for Add Event button",
     "Add Event button is not visible. Client cannot create events.",
     "High","Client"),

    # ── DOCUMENTS ──────────────────────────────────────────────────────────────
    ("CL-012","Documents","Access","Client can access Documents page",
     "1. Log in as Client user\n2. Navigate to Documents",
     "Documents page loads showing documents for the client's own appeals only.",
     "High","Client"),

    ("CL-013","Documents","Download","Client can download documents",
     "1. On Documents page or appeal detail\n2. Click Download on a document",
     "File downloads to local machine.",
     "High","Client"),

    ("CL-014","Documents","Upload","Client cannot upload documents",
     "1. Open appeal detail as Client user\n2. Look for Upload button in Documents section",
     "Upload button is not shown. Client cannot upload files.",
     "High","Client"),

    ("CL-015","Documents","Delete","Client cannot delete documents",
     "1. View document list as Client user",
     "Delete button is not visible on any document.",
     "High","Client"),

    # ── ACCESS CONTROL ─────────────────────────────────────────────────────────
    ("CL-016","Access Control","URL","Client cannot access /users via direct URL",
     "1. Log in as Client user\n2. Navigate directly to /users in browser URL bar",
     "Access denied or redirected to dashboard. Users page not rendered.",
     "High","Client"),

    ("CL-017","Access Control","URL","Client cannot access /clients via direct URL",
     "1. Log in as Client user\n2. Navigate directly to /clients",
     "Access denied or redirected. Clients page not rendered.",
     "High","Client"),

    ("CL-018","Access Control","URL","Client cannot access /logs via direct URL",
     "1. Log in as Client user\n2. Navigate directly to /logs",
     "Access denied or redirected. Logs page not rendered.",
     "High","Client"),

    ("CL-019","Access Control","URL","Client cannot access another org's appeal via direct URL",
     "1. Log in as Client user\n2. Copy URL of an appeal belonging to a different org\n3. Paste in browser",
     "Page shows not found or access denied. Other org's appeal data is not rendered.",
     "High","Client"),

    # ── SETTINGS ───────────────────────────────────────────────────────────────
    ("CL-020","Settings","Account","Client user can reset own password",
     "1. Use Forgot Password on login page\n2. Enter client email\n3. Open reset link\n4. Set new password",
     "New password is set. Client can log in with the new password.",
     "High","Client"),

    # ── GENERAL ────────────────────────────────────────────────────────────────
    ("CL-021","General","Session","Client session expires correctly",
     "1. Log in as Client user\n2. Leave session idle for extended period (if timeout configured)\n3. Try to perform action",
     "User is redirected to login page when session expires.",
     "Medium","Client"),

    ("CL-022","General","Responsive","Client portal usable on mobile",
     "1. Log in on a mobile browser or resize desktop to mobile width\n2. Navigate appeals list and detail",
     "Layout is responsive. Sidebar collapses, tables are scrollable, forms are usable on small screens.",
     "Medium","Client"),
]

# ─── Sheet builder ────────────────────────────────────────────────────────────

def add_sheet(wb, sheet_name, portal_label, accent_hex, section_bg_hex, tests):
    ws = wb.create_sheet(title=sheet_name)

    # Column widths
    col_widths = [10, 22, 20, 35, 50, 45, 10, 18, 12]
    col_letters = [get_column_letter(i+1) for i in range(len(col_widths))]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[col_letters[i]].width = w

    # ── Title row ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"AppealDesk — {portal_label} Functional Test Cases"
    title_cell.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    title_cell.fill = fill(accent_hex)
    title_cell.alignment = align("center", "center")
    title_cell.border = border_medium(accent_hex)
    ws.row_dimensions[1].height = 32

    # ── Sub-title ──────────────────────────────────────────────────────────────
    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = f"Portal: {portal_label}    |    Total Test Cases: {len(tests)}    |    Version: 1.0    |    April 2026"
    sub.font = Font(italic=True, color="FFFFFF", size=10, name="Calibri")
    sub.fill = fill(accent_hex)
    sub.alignment = align("center", "center")
    ws.row_dimensions[2].height = 18

    # ── Header row ─────────────────────────────────────────────────────────────
    headers = ["TC ID", "Section", "Module", "Test Case Name",
               "Test Steps", "Expected Result", "Priority", "Role", "Status"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        cell.fill = fill("374151")
        cell.alignment = align("center", "center")
        cell.border = border_thin("9CA3AF")
    ws.row_dimensions[3].height = 22

    # ── Data rows ──────────────────────────────────────────────────────────────
    priority_colors = {"High": "FEE2E2", "Medium": "FEF3C7", "Low": "D1FAE5"}
    priority_text   = {"High": "991B1B", "Medium": "92400E", "Low": "065F46"}

    current_section = None
    data_row = 4

    for tc in tests:
        tc_id, section, module, name, steps, expected, priority, role = tc

        # ── Section separator row ─────────────────────────────────────────────
        if section != current_section:
            current_section = section
            ws.merge_cells(f"A{data_row}:I{data_row}")
            sec_cell = ws.cell(row=data_row, column=1, value=f"  ▸  {section}")
            sec_cell.font = Font(bold=True, color=accent_hex, size=10, name="Calibri")
            sec_cell.fill = fill(section_bg_hex)
            sec_cell.alignment = align("left", "center", wrap=False)
            sec_cell.border = border_thin("D1D5DB")
            ws.row_dimensions[data_row].height = 18
            data_row += 1

        # ── Data row ──────────────────────────────────────────────────────────
        row_bg = C["white"] if (data_row % 2 == 0) else C["gray_row"]
        values = [tc_id, section, module, name, steps, expected, priority, role, ""]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.font = Font(color=C["blk"], size=10, name="Calibri")
            cell.border = border_thin()
            cell.alignment = align("left", "top", wrap=True)

            # TC ID bold
            if col == 1:
                cell.font = Font(bold=True, color=accent_hex, size=10, name="Calibri")
                cell.alignment = align("center", "top")

            # Priority badge colour
            elif col == 7:
                cell.fill = fill(priority_colors.get(priority, C["gray_row"]))
                cell.font = Font(bold=True, color=priority_text.get(priority, C["blk"]),
                                 size=10, name="Calibri")
                cell.alignment = align("center", "center")

            # Role
            elif col == 8:
                cell.alignment = align("center", "top")

            # Status (Pass/Fail/-)
            elif col == 9:
                cell.alignment = align("center", "center")

            # Background for other cols
            else:
                cell.fill = fill(row_bg)

        # Row height based on steps length
        lines = max(steps.count("\n"), expected.count("\n"), 1) + 2
        ws.row_dimensions[data_row].height = max(30, lines * 14)
        data_row += 1

    # ── Freeze header ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Auto-filter on headers ─────────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:I{data_row - 1}"

    return ws


# ─── Summary / Index sheet ────────────────────────────────────────────────────

def add_summary(wb):
    ws = wb.create_sheet(title="📋 Index", index=0)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 28

    # Title
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "AppealDesk — Functional Test Cases Index"
    t.font = Font(bold=True, color="FFFFFF", size=15, name="Calibri")
    t.fill = fill("1E3A5F")
    t.alignment = align("center", "center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:E2")
    s = ws["A2"]
    s.value = "MSSV & Co  |  Version 1.0  |  April 2026  |  Confidential"
    s.font = Font(italic=True, color="FFFFFF", size=10, name="Calibri")
    s.fill = fill("2D5491")
    s.alignment = align("center", "center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 10

    # Portal summary table header
    hdr_row = 4
    for col, h in enumerate(["", "Portal / Sheet", "Total TCs", "Priority Split", "Roles Covered"], start=1):
        c = ws.cell(row=hdr_row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.fill = fill("374151")
        c.alignment = align("center", "center")
        c.border = border_thin()
    ws.row_dimensions[hdr_row].height = 20

    portal_rows = [
        ("🔵", "Platform Portal", len(PLATFORM_TESTS),
         f"High: {sum(1 for t in PLATFORM_TESTS if t[6]=='High')}  Med: {sum(1 for t in PLATFORM_TESTS if t[6]=='Medium')}",
         "Super Admin"),
        ("🟢", "Service Provider Portal", len(SP_TESTS),
         f"High: {sum(1 for t in SP_TESTS if t[6]=='High')}  Med: {sum(1 for t in SP_TESTS if t[6]=='Medium')}",
         "SP Admin, SP Staff"),
        ("🟠", "Client Portal", len(CLIENT_TESTS),
         f"High: {sum(1 for t in CLIENT_TESTS if t[6]=='High')}  Med: {sum(1 for t in CLIENT_TESTS if t[6]=='Medium')}",
         "Client User"),
    ]

    bg_map = ["EFF6FF", "ECFDF5", "FFF7ED"]
    for i, (icon, name, count, split, roles) in enumerate(portal_rows):
        r = hdr_row + 1 + i
        data = [icon, name, count, split, roles]
        for col, val in enumerate(data, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill(bg_map[i])
            c.border = border_thin()
            c.alignment = align("center", "center")
            if col == 1:
                c.font = Font(size=14)
            elif col == 2:
                c.font = Font(bold=True, size=11, name="Calibri", color="1A1A2E")
                c.alignment = align("left", "center")
            elif col == 3:
                c.font = Font(bold=True, size=13, name="Calibri",
                              color=["1E3A5F","065F46","92400E"][i])
            else:
                c.font = Font(size=10, name="Calibri", color="374151")
        ws.row_dimensions[r].height = 24

    # Total row
    total_r = hdr_row + 4
    ws.merge_cells(f"A{total_r}:B{total_r}")
    tc = ws.cell(row=total_r, column=1, value="TOTAL")
    tc.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    tc.fill = fill("1E3A5F")
    tc.alignment = align("center", "center")
    tc.border = border_thin()
    total_count = len(PLATFORM_TESTS) + len(SP_TESTS) + len(CLIENT_TESTS)
    tval = ws.cell(row=total_r, column=3, value=total_count)
    tval.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    tval.fill = fill("1E3A5F")
    tval.alignment = align("center", "center")
    tval.border = border_thin()
    for col in [4, 5]:
        cx = ws.cell(row=total_r, column=col)
        cx.fill = fill("1E3A5F")
        cx.border = border_thin()
    ws.row_dimensions[total_r].height = 22

    # Gap
    ws.row_dimensions[total_r + 1].height = 10

    # Section breakdown
    sec_hdr = total_r + 2
    ws.merge_cells(f"A{sec_hdr}:E{sec_hdr}")
    sh = ws.cell(row=sec_hdr, column=1, value="Section Breakdown")
    sh.font = Font(bold=True, color="1E3A5F", size=11, name="Calibri")
    sh.fill = fill("E0E7FF")
    sh.alignment = align("left", "center")
    sh.border = border_thin()
    ws.row_dimensions[sec_hdr].height = 20

    # Collect sections
    all_tests = (
        [("Platform", t) for t in PLATFORM_TESTS] +
        [("SP", t) for t in SP_TESTS] +
        [("Client", t) for t in CLIENT_TESTS]
    )
    sections = {}
    for portal, t in all_tests:
        key = f"{portal} — {t[1]}"
        sections[key] = sections.get(key, 0) + 1

    for j, (sec, cnt) in enumerate(sections.items()):
        r = sec_hdr + 1 + j
        ws.merge_cells(f"A{r}:D{r}")
        sc = ws.cell(row=r, column=1, value=f"  {sec}")
        sc.font = Font(size=10, name="Calibri", color="374151")
        sc.fill = fill("F9FAFB" if j % 2 == 0 else "FFFFFF")
        sc.alignment = align("left", "center")
        sc.border = border_thin()
        cc = ws.cell(row=r, column=5, value=cnt)
        cc.font = Font(bold=True, size=10, name="Calibri", color="1E3A5F")
        cc.fill = fill("F9FAFB" if j % 2 == 0 else "FFFFFF")
        cc.alignment = align("center", "center")
        cc.border = border_thin()
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = "A4"

# ─── Build workbook ───────────────────────────────────────────────────────────

add_summary(wb)
add_sheet(wb, "🔵 Platform Portal",  "Platform Portal",         C["navy"],  "EFF6FF", PLATFORM_TESTS)
add_sheet(wb, "🟢 SP Portal",        "Service Provider Portal", C["green"], "ECFDF5", SP_TESTS)
add_sheet(wb, "🟠 Client Portal",    "Client Portal",           C["orange"],"FFF7ED", CLIENT_TESTS)

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Total TCs: {len(PLATFORM_TESTS) + len(SP_TESTS) + len(CLIENT_TESTS)}")
